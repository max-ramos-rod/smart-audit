"""Unit tests for the form import parser (CSV and Excel)."""
import io

import pytest
from fastapi import HTTPException

from app.modules.forms.importer import parse_csv, parse_excel, parse_import_file


def _make_csv(
    *rows: str, header: str = "label,tipo,instrucao,obrigatorio,peso,opcoes,permite_na"
) -> bytes:
    lines = [header] + list(rows)
    return "\n".join(lines).encode("utf-8")


class TestParseCsv:
    def test_boolean_field_parsed(self):
        content = _make_csv("Extintor presente?,sim_nao,,sim,,, ")
        req = parse_csv(content, "Teste")
        assert len(req.fields) == 1
        f = req.fields[0]
        assert f.field_type == "boolean"
        assert f.label == "Extintor presente?"
        assert f.required is True

    def test_section_creates_divider(self):
        content = _make_csv("Bloco A,secao", "Item,sim_nao")
        req = parse_csv(content, "Teste")
        assert req.fields[0].field_type == "section"
        assert req.fields[1].field_type == "boolean"

    def test_instruction_populated(self):
        content = _make_csv("Campo,texto,Verificar com atenção,sim")
        req = parse_csv(content, "Teste")
        assert req.fields[0].instruction == "Verificar com atenção"

    def test_instruction_none_when_empty(self):
        content = _make_csv("Campo,texto,,sim")
        req = parse_csv(content, "Teste")
        assert req.fields[0].instruction is None

    def test_select_options_parsed(self):
        content = _make_csv("Estado,selecao,,,, Bom;Regular;Ruim,")
        req = parse_csv(content, "Teste")
        assert req.fields[0].field_type == "select"
        assert req.fields[0].config_json["options"] == ["Bom", "Regular", "Ruim"]

    def test_select_without_opcoes_raises(self):
        content = _make_csv("Estado,selecao")
        with pytest.raises(HTTPException) as exc:
            parse_csv(content, "Teste")
        assert exc.value.status_code == 422

    def test_boolean_with_weight(self):
        content = _make_csv("Crítico,sim_nao,,,3,,")
        req = parse_csv(content, "Teste")
        assert req.fields[0].config_json.get("weight") == 3.0

    def test_boolean_default_weight_omitted(self):
        content = _make_csv("Normal,sim_nao,,,1,,")
        req = parse_csv(content, "Teste")
        assert "weight" not in req.fields[0].config_json

    def test_permite_na_sets_allow_na(self):
        content = _make_csv("Campo,sim_nao,,,,, sim")
        req = parse_csv(content, "Teste")
        assert req.fields[0].config_json.get("allow_na") is True

    def test_not_required_field(self):
        content = _make_csv("Opcional,texto,,nao")
        req = parse_csv(content, "Teste")
        assert req.fields[0].required is False

    def test_form_name_set(self):
        content = _make_csv("Campo,texto")
        req = parse_csv(content, "Meu Formulário")
        assert req.name == "Meu Formulário"

    def test_duplicate_keys_get_suffix(self):
        content = _make_csv("Item,sim_nao", "Item,sim_nao")
        req = parse_csv(content, "Teste")
        keys = [f.key for f in req.fields]
        assert len(set(keys)) == 2

    def test_invalid_tipo_raises(self):
        content = _make_csv("Campo,invalido")
        with pytest.raises(HTTPException) as exc:
            parse_csv(content, "Teste")
        assert exc.value.status_code == 422

    def test_empty_rows_skipped(self):
        content = _make_csv("Campo,texto", ",")
        req = parse_csv(content, "Teste")
        assert len(req.fields) == 1

    def test_missing_label_raises(self):
        content = _make_csv(",sim_nao")
        with pytest.raises(HTTPException) as exc:
            parse_csv(content, "Teste")
        assert exc.value.status_code == 422

    def test_missing_required_column_raises(self):
        content = b"label\nCampo"
        with pytest.raises(HTTPException) as exc:
            parse_csv(content, "Teste")
        assert exc.value.status_code == 422

    def test_utf8_bom_handled(self):
        content = b"\xef\xbb\xbflabel,tipo\nCampo,texto"
        req = parse_csv(content, "Teste")
        assert req.fields[0].label == "Campo"

    def test_alternative_tipo_names_accepted(self):
        mapping = [
            ("texto", "text"),
            ("numero", "number"),
            ("data", "date"),
            ("boolean", "boolean"),
            ("text", "text"),
        ]
        for tipo, expected in mapping:
            content = _make_csv(f"C,{tipo}")
            req = parse_csv(content, "T")
            assert req.fields[0].field_type == expected, f"Falhou para tipo={tipo}"


class TestParseImportFile:
    def test_csv_extension_dispatched(self):
        content = _make_csv("Campo,texto")
        req = parse_import_file(content, "checklist.csv", None)
        assert req.fields[0].field_type == "text"

    def test_name_from_filename_when_not_provided(self):
        content = _make_csv("Campo,texto")
        req = parse_import_file(content, "meu_checklist.csv", None)
        assert req.name == "Meu Checklist"

    def test_explicit_name_overrides_filename(self):
        content = _make_csv("Campo,texto")
        req = parse_import_file(content, "arquivo.csv", "Nome Explícito")
        assert req.name == "Nome Explícito"

    def test_unsupported_extension_raises(self):
        with pytest.raises(HTTPException) as exc:
            parse_import_file(b"data", "file.pdf", None)
        assert exc.value.status_code == 415

    def test_xlsx_extension_dispatched(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["label", "tipo", "instrucao", "obrigatorio", "peso", "opcoes", "permite_na"])
        ws.append(["Campo", "texto", "Instrução", "sim", "", "", ""])
        buf = io.BytesIO()
        wb.save(buf)
        req = parse_import_file(buf.getvalue(), "form.xlsx", "Excel Form")
        assert req.fields[0].field_type == "text"
        assert req.fields[0].instruction == "Instrução"


class TestParseExcel:
    def _make_xlsx(self, rows: list[list]) -> bytes:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["label", "tipo", "instrucao", "obrigatorio", "peso", "opcoes", "permite_na"])
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_boolean_parsed(self):
        content = self._make_xlsx([["Extintor", "sim_nao", "", "sim", "", "", ""]])
        req = parse_excel(content, "Teste")
        assert req.fields[0].field_type == "boolean"

    def test_section_parsed(self):
        content = self._make_xlsx([["Bloco A", "secao", "", "", "", "", ""],
                                   ["Item", "sim_nao", "", "sim", "", "", ""]])
        req = parse_excel(content, "Teste")
        assert req.fields[0].field_type == "section"
        assert req.fields[1].field_type == "boolean"

    def test_select_with_options(self):
        content = self._make_xlsx([["Estado", "selecao", "", "sim", "", "Bom;Regular;Ruim", ""]])
        req = parse_excel(content, "Teste")
        assert req.fields[0].config_json["options"] == ["Bom", "Regular", "Ruim"]

    def test_missing_required_column_raises(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["label"])
        ws.append(["Campo"])
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(HTTPException) as exc:
            parse_excel(buf.getvalue(), "Teste")
        assert exc.value.status_code == 422
